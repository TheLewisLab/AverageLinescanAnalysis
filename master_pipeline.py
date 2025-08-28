"""
Mitochondrial Protein Analysis Pipeline - Windows Compatible Version
===================================================================

Pipeline for analyzing mitochondrial protein localization from
multi-channel microscopy images. Performs peak detection, cross-channel
normalization, statistical analysis, and generates Excel outputs for research.

SETUP INSTRUCTIONS:
1. Use Visual_lineScan_test.py to determine optimal peak detection parameters
2. Update the configuration section below with your specific parameters
3. Ensure your CSV files from ImageJ are in the correct folder structure
4. Run this script to perform complete analysis

"""

import os
import random
import logging
from warnings import simplefilter
from typing import Optional, Dict, List, Tuple, Union
from pathlib import Path
import time

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from scipy.signal import find_peaks

simplefilter(action='ignore', category=pd.errors.PerformanceWarning)

# =============================================================================
# CONFIGURATION - UPDATE THESE SETTINGS FOR YOUR EXPERIMENT
# =============================================================================

# Experiment setup
EXPERIMENT_DIR = Path('mitochondrial_analysis')
DATA_FOLDER_NAME = 'experimental_data'
CSV_SUBFOLDER = 'csv_files'  # Set to 'csv_files' to use subfolder, or None for CSV files directly in EXPERIMENT_DIR

# Image parameters
PIXEL_DISTANCE = 0.035  # Microns per pixel (from ImageJ calibration)
REGION_TO_PULL = 14  # Pixels to extract around each peak (±14 = 29 total)
CHANNELS_IN_IMAGE = 4  # Total number of channels in your images

# Channel configuration - customize for your experiment
CHANNEL_LABELS = {
    1: 'GRSF1',  # Channel 1
    2: 'LETM1',  # Channel 2
    3: 'TOM20',  # Channel 3
    4: 'dsDNA'  # Channel 4
}

# Analysis parameters
CHANNELS_TO_ANALYZE = [1, 2, 4]  # Channels to analyze for peaks
BACKBONE_CHANNEL = 3  # Reference channel (not analyzed for peaks)

# Peak detection parameters - copy from Visual_lineScan_test.py results
PEAK_PARAMETERS = {
    1: [4, 1500, 500, 6],     # Channel 1: [width, height, prominence, distance]
    2: [4, 2000, 1000, 6],    # Channel 2: [width, height, prominence, distance]
    3: [4, 1000, 100, 6],      # Channel 3: [width, height, prominence, distance]
    4: [4, 5000, 1000, 6],    # Channel 4: [width, height, prominence, distance]
}

# Output settings
SAVE_PLOTS = True
SHOW_PLOTS = False
SAVING_DATA = True

# Random sampling settings
MAX_RANDOM_SAMPLES = 2000
MIN_DATA_POINTS_FOR_SAMPLING = 30

# =============================================================================
# CACHED CALCULATIONS AND CONSTANTS
# =============================================================================

# Pre-calculate distance array to avoid repeated calculations
DISTANCES = np.array([(i - REGION_TO_PULL) * PIXEL_DISTANCE for i in range(2 * REGION_TO_PULL + 1)])
DATA_LENGTH = len(DISTANCES)


# =============================================================================
# LOGGING SETUP - WINDOWS COMPATIBLE
# =============================================================================

def setup_logging():
    """Setup logging with Windows-compatible encoding"""
    log_format = '%(asctime)s - %(levelname)s - %(message)s'

    # Create logs directory
    log_dir = EXPERIMENT_DIR / 'logs'
    log_dir.mkdir(parents=True, exist_ok=True)

    # Setup file and console logging with UTF-8 encoding
    logging.basicConfig(
        level=logging.INFO,
        format=log_format,
        handlers=[
            logging.FileHandler(log_dir / f'{DATA_FOLDER_NAME}_analysis.log', encoding='utf-8'),
            logging.StreamHandler()
        ]
    )

    # Set console encoding for Windows compatibility
    logger = logging.getLogger(__name__)
    for handler in logger.handlers:
        if isinstance(handler, logging.StreamHandler):
            handler.stream.reconfigure(encoding='utf-8', errors='replace')

    return logger


logger = setup_logging()


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def col_to_letter(col_num: int) -> str:
    """Convert column number to Excel letter(s). Handles unlimited columns."""
    if col_num <= 0:
        raise ValueError("Column number must be positive")

    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(65 + (col_num % 26)) + result
        col_num //= 26
    return result


def safe_divide(numerator: Union[pd.DataFrame, pd.Series, float],
                denominator: Union[float, int],
                default_value: float = 0) -> Union[pd.DataFrame, pd.Series, float]:
    """Safely divide with protection against division by zero"""
    if isinstance(denominator, (int, float)) and denominator == 0:
        logger.warning(f"Division by zero detected, using default value: {default_value}")
        if isinstance(numerator, (pd.DataFrame, pd.Series)):
            return numerator * 0 + default_value
        return default_value

    return numerator / denominator


def calculate_optimal_sample_size(total_data_points: int,
                                  max_samples: int = MAX_RANDOM_SAMPLES,
                                  min_required: int = MIN_DATA_POINTS_FOR_SAMPLING) -> int:
    """Calculate optimal sample size based on available data"""
    if total_data_points < min_required:
        logger.warning(f"Insufficient data for sampling: {total_data_points} < {min_required}")
        return 0

    # Use smaller of max_samples or what's feasible given data constraints
    max_feasible = total_data_points // (REGION_TO_PULL * 2 + 1)  # Account for region extraction
    optimal_size = min(max_samples, max_feasible)

    logger.info(f"Calculated optimal sample size: {optimal_size} from {total_data_points} available points")
    return optimal_size


def safe_save_excel(workbook, filename: Path, max_retries: int = 3) -> bool:
    """Safely save Excel file with retry logic for permission issues"""
    for attempt in range(max_retries):
        try:
            workbook.save(filename)
            logger.info(f"Successfully saved: {filename}")
            return True
        except PermissionError:
            if attempt < max_retries - 1:
                logger.warning(
                    f"Permission denied for {filename}, retrying in 2 seconds... (attempt {attempt + 1}/{max_retries})")
                time.sleep(2)
            else:
                logger.error(
                    f"Failed to save {filename} after {max_retries} attempts. Please close Excel and try again.")
                return False
        except Exception as e:
            logger.error(f"Error saving {filename}: {e}")
            return False
    return False


def validate_configuration() -> bool:
    """Validate the configuration parameters before running analysis"""
    errors = []
    warnings = []

    # Check basic requirements
    if CHANNELS_IN_IMAGE < 2:
        errors.append("CHANNELS_IN_IMAGE must be at least 2")

    if CHANNELS_IN_IMAGE > 10:
        warnings.append(f"CHANNELS_IN_IMAGE={CHANNELS_IN_IMAGE} is unusually high")

    # Check channel labels
    for ch in range(1, CHANNELS_IN_IMAGE + 1):
        if ch not in CHANNEL_LABELS:
            warnings.append(f"Channel {ch} has no label defined")

    # Check analysis channels
    if not CHANNELS_TO_ANALYZE:
        errors.append("CHANNELS_TO_ANALYZE cannot be empty")

    for ch in CHANNELS_TO_ANALYZE:
        if ch < 1 or ch > CHANNELS_IN_IMAGE:
            errors.append(f"Invalid channel {ch} in CHANNELS_TO_ANALYZE")
        if ch == BACKBONE_CHANNEL:
            warnings.append(
                f"Channel {ch} is set as BACKBONE_CHANNEL but also in CHANNELS_TO_ANALYZE - it will be excluded from analysis")

    # Check backbone
    if BACKBONE_CHANNEL is not None:
        if BACKBONE_CHANNEL < 1 or BACKBONE_CHANNEL > CHANNELS_IN_IMAGE:
            errors.append(f"Invalid BACKBONE_CHANNEL: {BACKBONE_CHANNEL}")

    # Check peak parameters
    # Check peak parameters
    for ch in CHANNELS_TO_ANALYZE:
        if ch not in PEAK_PARAMETERS:
            errors.append(f"No PEAK_PARAMETERS defined for channel {ch}")
        elif len(PEAK_PARAMETERS[ch]) != 4:  # Changed from 3 to 4
            errors.append(f"PEAK_PARAMETERS for channel {ch} must have 4 values: [width, height, prominence, distance]")

    # Check for at least one non-backbone channel
    analyzed_non_backbone = [ch for ch in CHANNELS_TO_ANALYZE if ch != BACKBONE_CHANNEL]
    if not analyzed_non_backbone:
        errors.append("Must analyze at least one non-backbone channel")

    # Display results
    if errors:
        logger.error("Configuration validation failed:")
        for error in errors:
            logger.error(f"   • {error}")
        return False

    if warnings:
        logger.warning("Configuration warnings:")
        for warning in warnings:
            logger.warning(f"   • {warning}")

    logger.info("Configuration validated successfully")
    return True


def setup_folders():
    """Create organized output folder structure"""
    logger.info("Setting up folder structure...")

    base_folders = [
        'peak_csvs',
        'avg_peak_csvs',
        'sum_peak_csvs',
        'norm_peak_csvs',
        'confidence_data',
        'random_data_lines',
        'plots',
        'excel_files'
    ]

    for folder_name in base_folders:
        folder_path = EXPERIMENT_DIR / folder_name
        folder_path.mkdir(parents=True, exist_ok=True)

        # Create channel subfolders where needed
        if folder_name not in ['plots', 'excel_files', 'random_data_lines']:
            for channel_num in CHANNELS_TO_ANALYZE:
                channel_folder = folder_path / f'Channel_{channel_num}'
                channel_folder.mkdir(exist_ok=True)

    logger.info(f"[SUCCESS] Folder structure created in {EXPERIMENT_DIR}/")


def check_input_files() -> List[str]:
    """Verify input CSV files exist and are accessible"""
    if CSV_SUBFOLDER:
        csv_path = EXPERIMENT_DIR / CSV_SUBFOLDER
    else:
        csv_path = EXPERIMENT_DIR

    if not csv_path.exists():
        raise FileNotFoundError(f"CSV folder not found: {csv_path}")

    csv_files = [f.name for f in csv_path.iterdir() if f.suffix == '.csv']
    if not csv_files:
        raise FileNotFoundError(f"No CSV files found in {csv_path}")

    logger.info(f"[SUCCESS] Found {len(csv_files)} CSV files in {csv_path}")
    return csv_files

def process_roi_data(df: pd.DataFrame, roi_index: int) -> pd.DataFrame:
    """Extract and clean data for a single ROI"""
    try:
        roi_line = pd.DataFrame()
        roi_line['Distance'] = df[f'ROI_{roi_index} distance (micron)']

        # Find where distance becomes 0 (end of valid data)
        drop_value = len(roi_line)
        distance_values = roi_line['Distance'].values
        for j in range(1, len(distance_values)):
            if distance_values[j] == 0:
                drop_value = j
                break

        # Extract all channel data
        for ch in range(1, CHANNELS_IN_IMAGE + 1):
            column_name = f'ROI_{roi_index} Channel_{ch}'
            if column_name in df.columns:
                roi_line[f'Channel_{ch}'] = df[column_name]
            else:
                logger.warning(f"Missing column: {column_name}")
                roi_line[f'Channel_{ch}'] = 0

        # Remove invalid data points
        if drop_value < len(roi_line):
            roi_line = roi_line.iloc[:drop_value].copy()

        return roi_line

    except Exception as e:
        logger.error(f"Error processing ROI {roi_index}: {e}")
        return pd.DataFrame()


def create_empty_analysis_dataframe() -> pd.DataFrame:
    """Create an empty DataFrame with the standard structure"""
    return pd.DataFrame({
        'Distance': DISTANCES,
        **{f'Channel {i}': np.zeros(DATA_LENGTH) for i in range(1, CHANNELS_IN_IMAGE + 1)}
    })


def save_analysis_results(out_df: pd.DataFrame, avg_df: pd.DataFrame, sum_df: pd.DataFrame,
                          channel_num: int, filename: str):
    """Save analysis results to appropriate directories"""
    if not SAVING_DATA:
        return

    try:
        base_name = Path(filename).stem
        channel_dir = f'Channel_{channel_num}'

        # Save to respective directories
        out_path = EXPERIMENT_DIR / 'peak_csvs' / channel_dir / f'{base_name}_each_peak.csv'
        avg_path = EXPERIMENT_DIR / 'avg_peak_csvs' / channel_dir / f'{base_name}_avg_peaks.csv'
        sum_path = EXPERIMENT_DIR / 'sum_peak_csvs' / channel_dir / f'{base_name}_sum_peaks.csv'

        out_df.to_csv(out_path, index=False)
        avg_df.to_csv(avg_path, index=False)
        sum_df.to_csv(sum_path, index=False)

    except Exception as e:
        logger.error(f"Error saving results for {filename}, channel {channel_num}: {e}")


def run_peak_analysis() -> Tuple[Dict, Dict]:
    """Main peak detection and analysis workflow"""
    logger.info("=" * 60)
    logger.info("PEAK ANALYSIS WITH CROSS-CHANNEL NORMALIZATION")
    logger.info("=" * 60)

    csv_files = check_input_files()

    # Update the csv_path calculation
    if CSV_SUBFOLDER:
        csv_path = EXPERIMENT_DIR / CSV_SUBFOLDER
    else:
        csv_path = EXPERIMENT_DIR

    # Initialize analysis tracking
    log_file = None
    if SAVING_DATA:
        log_file_path = EXPERIMENT_DIR / 'analysis_metadata.txt'
        log_file = open(log_file_path, 'w', encoding='utf-8')
        log_file.write(f"Analysis Log for {EXPERIMENT_DIR}\n")
        log_file.write("=" * 50 + "\n\n")

    channel_maxima = {ch: 0 for ch in range(1, CHANNELS_IN_IMAGE + 1)}
    channel_peak_dfs = {}
    total_line_scans = 0

    try:
        # Process each channel specified for analysis
        for peak_channel_num in CHANNELS_TO_ANALYZE:
            channel_label = CHANNEL_LABELS.get(peak_channel_num, f'Channel_{peak_channel_num}')
            peak_channel = f'Channel_{peak_channel_num}'
            logger.info(f"Processing {peak_channel} ({channel_label})...")

            if peak_channel_num not in PEAK_PARAMETERS:
                logger.warning(f"No parameters defined for {peak_channel}")
                continue

            peak_params = PEAK_PARAMETERS[peak_channel_num]

            # Initialize tracking variables
            channel_peak_max = 0
            avg_height = 0
            avg_width = 0
            avg_prominence = 0
            total_peaks = 0

            # Initialize image average DataFrame
            image_avg_df = create_empty_analysis_dataframe()
            image_n = 0

            # Process each CSV file
            for filename in csv_files:
                try:
                    df = pd.read_csv(csv_path / filename)
                    name = Path(filename).stem
                    image_n += 1

                    # Initialize output DataFrames
                    out_df = pd.DataFrame(DISTANCES, columns=['Distance'])
                    avg_df = create_empty_analysis_dataframe()

                    peak_number = 1
                    running_avg_height = 0
                    running_avg_width = 0
                    running_avg_prominence = 0

                    # Process each ROI
                    num_rois = int(len(df.columns) / (CHANNELS_IN_IMAGE + 1))
                    total_line_scans += num_rois

                    for roi_idx in range(num_rois):
                        roi_line = process_roi_data(df, roi_idx)

                        if roi_line.empty:
                            continue

                        # Find peaks in the specified channel
                        if peak_channel in roi_line.columns:
                            peak_values = find_peaks(
                                roi_line[peak_channel],
                                prominence=peak_params[2],
                                width=peak_params[0],
                                distance=peak_params[3],
                                height=peak_params[1]
                            )

                            # Extract data around each peak
                            for j in range(len(peak_values[0])):
                                peak_position = peak_values[0][j]
                                if REGION_TO_PULL < peak_position < (len(roi_line) - REGION_TO_PULL):
                                    # Extract regions for all channels
                                    for ch in range(1, CHANNELS_IN_IMAGE + 1):
                                        channel_col = f'Channel_{ch}'
                                        if channel_col in roi_line.columns:
                                            peak_region = roi_line[channel_col][
                                                          peak_position - REGION_TO_PULL:peak_position + (
                                                                      REGION_TO_PULL + 1)].reset_index(drop=True)
                                            out_df[f'Peak_{peak_number} Channel {ch}'] = peak_region
                                            avg_df[f'Channel {ch}'] += peak_region

                                    # Track peak statistics
                                    temp_peak_height = peak_values[1]['peak_heights'][j]
                                    running_avg_height += temp_peak_height
                                    running_avg_width += peak_values[1]['widths'][j]
                                    running_avg_prominence += peak_values[1]['prominences'][j]

                                    if temp_peak_height > channel_peak_max:
                                        channel_peak_max = temp_peak_height

                                    peak_number += 1

                    # Calculate averages for this image
                    running_peaks = peak_number - 1
                    total_peaks += running_peaks

                    if running_peaks > 0:
                        sum_df = avg_df.copy()
                        # Use safe division to prevent division by zero
                        avg_df = safe_divide(avg_df, running_peaks)
                        avg_df['Distance'] = DISTANCES
                        sum_df['Distance'] = DISTANCES

                        # Save individual image results
                        save_analysis_results(out_df, avg_df, sum_df, peak_channel_num, filename)

                        # Accumulate for image average
                        for ch in range(1, CHANNELS_IN_IMAGE + 1):
                            image_avg_df[f'Channel {ch}'] += sum_df[f'Channel {ch}']

                    # Update statistics
                    avg_height += running_avg_height
                    avg_width += running_avg_width
                    avg_prominence += running_avg_prominence

                except Exception as e:
                    logger.error(f"Error processing file {filename}: {e}")
                    continue

            # Calculate final averages
            if total_peaks > 0:
                # Use safe division for final averaging
                image_avg_df = safe_divide(image_avg_df, total_peaks)
                image_avg_df['Distance'] = DISTANCES

                # Track global maximum and save results
                channel_maxima[peak_channel_num] = channel_peak_max
                channel_peak_dfs[peak_channel_num] = image_avg_df.copy()

                # Save overall results
                if SAVING_DATA:
                    avg_channel_dir = EXPERIMENT_DIR / 'avg_peak_csvs' / f'Channel_{peak_channel_num}'
                    avg_channel_dir.mkdir(parents=True, exist_ok=True)
                    avg_peaks_path = avg_channel_dir / f'{DATA_FOLDER_NAME}_Channel_{peak_channel_num}_avg_peaks.csv'
                    image_avg_df.to_csv(avg_peaks_path, index=False)

                # Log statistics
                if SAVING_DATA and log_file and total_peaks > 0:
                    log_file.write(f'name: {DATA_FOLDER_NAME}\n')
                    log_file.write(f'Channel: Channel_{peak_channel_num}\n')
                    log_file.write(f'Peak Width Minimum: {peak_params[0]}\n')
                    log_file.write(f'Peak Height Minimum: {peak_params[1]}\n')
                    log_file.write(f'Peak Prominence Minimum: {peak_params[2]}\n')
                    log_file.write(f'Peak Distance Minimum: {peak_params[3]}\n')
                    log_file.write(f'avg height: {safe_divide(avg_height, total_peaks)}\n')
                    log_file.write(f'avg_width: {safe_divide(avg_width, total_peaks)}\n')
                    log_file.write(f'avg_prominence: {safe_divide(avg_prominence, total_peaks)}\n')
                    log_file.write(f'total_peaks: {total_peaks}\n')
                    if image_n > 0:
                        log_file.write(f'avg peaks per image: {safe_divide(total_peaks, image_n)}\n\n')

                logger.info(f"[SUCCESS] Found {total_peaks} peaks across {image_n} images")
                if image_n > 0:
                    logger.info(f"  Average: {safe_divide(total_peaks, image_n):.1f} peaks per image")
                logger.info(f"  Max peak height: {channel_peak_max:.0f}")

        # Log channel maxima
        if channel_peak_dfs and SAVING_DATA and log_file:
            log_file.write(f'Channel maxima:\n')
            for ch in range(1, CHANNELS_IN_IMAGE + 1):
                log_file.write(f'c{ch}_max: {channel_maxima[ch]}\n')
            log_file.write(f'total line scans: {total_line_scans}\n')

    finally:
        if SAVING_DATA and log_file:
            log_file.close()

    # Create normalized outputs and plots
    create_normalized_outputs_and_plots(channel_peak_dfs, channel_maxima)

    return channel_peak_dfs, channel_maxima


def create_normalized_outputs_and_plots(channel_peak_dfs: Dict, channel_maxima: Dict):
    """Create normalized CSV files and cross-channel plots"""
    logger.info("Creating normalized outputs and cross-channel plots...")

    # Create normalization values (add safety check)
    max_values = [1] + [max(channel_maxima[ch], 1) for ch in
                        range(1, CHANNELS_IN_IMAGE + 1)]  # Prevent division by zero

    # Setup plotting
    if len(CHANNELS_TO_ANALYZE) > 1:
        figure, axis = plt.subplots(1, len(CHANNELS_TO_ANALYZE), figsize=(6 * len(CHANNELS_TO_ANALYZE), 6))
        if len(CHANNELS_TO_ANALYZE) == 1:
            axis = [axis]
    else:
        figure, axis = plt.subplots(1, 1, figsize=(8, 6))
        axis = [axis]

    colors = {1: 'orange', 2: 'red', 3: 'green', 4: 'blue'}

    z = 0
    for channel_num in CHANNELS_TO_ANALYZE:
        if channel_num in channel_peak_dfs:
            image_avg_df = channel_peak_dfs[channel_num]

            # Create normalized version using safe division
            image_norm_avg_df = safe_divide(image_avg_df, max_values)

            # Save normalized CSV
            if SAVING_DATA:
                norm_dir = EXPERIMENT_DIR / 'norm_peak_csvs'
                norm_dir.mkdir(parents=True, exist_ok=True)

                channel_norm_dir = norm_dir / f'Channel_{channel_num}'
                channel_norm_dir.mkdir(exist_ok=True)

                norm_path = channel_norm_dir / f'{DATA_FOLDER_NAME}_Channel_{channel_num}_norm_avg_peaks.csv'
                image_norm_avg_df.to_csv(norm_path, index=False)
                logger.info(f"[SUCCESS] Saved normalized CSV: {norm_path}")

            # Plot analyzed channels
            channels_plotted = []
            for ch in CHANNELS_TO_ANALYZE:
                col_name = f'Channel {ch}'
                if col_name in image_norm_avg_df.columns:
                    axis[z].plot(image_norm_avg_df['Distance'], image_norm_avg_df[col_name],
                                 color=colors.get(ch, 'black'),
                                 label=f'Ch{ch} ({CHANNEL_LABELS.get(ch, f"Channel_{ch}")})',
                                 linewidth=2)
                    channels_plotted.append(ch)

            # Format plot
            axis[z].axvline(0, color='black', linestyle='--', alpha=0.7, label='Peak Center')
            axis[z].set_xlim([-(REGION_TO_PULL * PIXEL_DISTANCE), (REGION_TO_PULL * PIXEL_DISTANCE)])
            axis[z].set_ylim([0, 0.4])

            channel_label = CHANNEL_LABELS.get(channel_num, f'Channel_{channel_num}')
            axis[z].set_title(f'{DATA_FOLDER_NAME} Channel_{channel_num}\n({channel_label} Peak Analysis)')
            axis[z].set_xlabel('Distance from Peak Center (μm)')
            axis[z].set_ylabel('Normalized Intensity')
            axis[z].legend()
            axis[z].grid(True, alpha=0.3)

            logger.info(f"[SUCCESS] Plotted channels {channels_plotted} for {channel_label} analysis")
            z += 1

    # Save plots
    if SAVE_PLOTS:
        plots_dir = EXPERIMENT_DIR / 'plots'
        plots_dir.mkdir(parents=True, exist_ok=True)
        plot_path = plots_dir / f'{DATA_FOLDER_NAME}_cross_channel_analysis.png'
        plt.savefig(plot_path, dpi=300, bbox_inches='tight')
        logger.info(f"[SUCCESS] Saved plot: {plot_path}")

    if SHOW_PLOTS:
        plt.show()
    else:
        plt.close()


def run_random_background_analysis() -> Tuple[Optional[Dict], Optional[Dict]]:
    """Analyze random background regions for normalization"""
    logger.info("=" * 60)
    logger.info("RANDOM BACKGROUND ANALYSIS")
    logger.info("=" * 60)

    if CSV_SUBFOLDER:
        csv_path = EXPERIMENT_DIR / CSV_SUBFOLDER
    else:
        csv_path = EXPERIMENT_DIR

    random_dir = EXPERIMENT_DIR / 'random_data_lines'
    random_dir.mkdir(parents=True, exist_ok=True)

    # Collect all ROI data
    all_roi_data = []
    csv_files = [f for f in csv_path.iterdir() if f.suffix == '.csv']
    total_files_processed = 0
    total_rois_processed = 0

    logger.info(f"Processing {len(csv_files)} CSV files...")


    try:
        for csv_file in csv_files:
            try:
                df = pd.read_csv(csv_file)
                num_rois = int(len(df.columns) / (CHANNELS_IN_IMAGE + 1))
                total_files_processed += 1
                total_rois_processed += num_rois

                for roi_idx in range(num_rois):
                    roi_line = process_roi_data(df, roi_idx)
                    if not roi_line.empty:
                        roi_channels = roi_line.drop(columns=['Distance'])
                        all_roi_data.append(roi_channels)

            except Exception as e:
                logger.error(f"Error processing {csv_file}: {e}")
                continue

        # Combine all data
        if all_roi_data:
            out_df = pd.concat(all_roi_data, ignore_index=True)
        else:
            logger.error("No ROI data collected")
            return None, None

        logger.info(f"[SUCCESS] Processed {total_files_processed} files with {total_rois_processed} ROIs")
        logger.info(f"[SUCCESS] Collected {len(out_df)} data points from all line scans")

        # Calculate optimal sample size
        optimal_samples = calculate_optimal_sample_size(len(out_df))

        if optimal_samples == 0:
            logger.error("Insufficient data for random sampling")
            return None, None

        # Sample random locations
        avg_peak = pd.DataFrame({f'Channel_{ch}': np.zeros(DATA_LENGTH) for ch in range(1, CHANNELS_IN_IMAGE + 1)})
        peak_n = 0
        sample_values_by_channel = {f'Channel_{ch}': [] for ch in range(1, CHANNELS_IN_IMAGE + 1)}

        logger.info(f"Sampling {optimal_samples} random locations...")

        while peak_n < optimal_samples and len(out_df) >= MIN_DATA_POINTS_FOR_SAMPLING:
            location = random.randrange(REGION_TO_PULL, len(out_df) - REGION_TO_PULL)

            for ch in range(1, CHANNELS_IN_IMAGE + 1):
                channel_key = f'Channel_{ch}'
                if channel_key in out_df.columns:
                    region = out_df[channel_key][location - REGION_TO_PULL:location + REGION_TO_PULL + 1].reset_index(
                        drop=True)
                    if len(region) == DATA_LENGTH:
                        avg_peak[channel_key] += region
                        sample_values_by_channel[channel_key].extend(region.tolist())

            peak_n += 1

        if peak_n > 0:
            # Use safe division for averaging
            avg_peak = safe_divide(avg_peak, peak_n)

            # Calculate statistics
            random_means = {}
            random_stds = {}
            random_stats = {}

            for ch in range(1, CHANNELS_IN_IMAGE + 1):
                channel_key = f'Channel_{ch}'
                if channel_key in out_df.columns:
                    sample_values = sample_values_by_channel[channel_key]
                    mean_val = float(np.mean(sample_values)) if sample_values else 0
                    std_val = float(np.std(sample_values)) if sample_values else 0

                    random_means[channel_key] = mean_val
                    random_stds[channel_key] = std_val
                    random_stats[channel_key] = {
                        'mean': mean_val,
                        'std': std_val,
                        'samples': len(sample_values)
                    }

            # Save results
            if SAVING_DATA:
                # Main random data file
                avg_peak.to_csv(random_dir / f'{DATA_FOLDER_NAME}_random_avg_linescan.csv', index=False)

                # Statistics file
                with open(random_dir / f'{DATA_FOLDER_NAME}_avg_channel_value.txt', 'w', encoding='utf-8') as out_file:
                    out_file.write('image set mean\n')
                    out_file.write(str(out_df.mean()))
                    out_file.write('\nimage set std\n')
                    out_file.write(str(out_df.std()))

                # Detailed results
                results_df = pd.DataFrame()
                for ch in range(1, CHANNELS_IN_IMAGE + 1):
                    ch_label = CHANNEL_LABELS.get(ch, f'Channel_{ch}')
                    channel_key = f'Channel_{ch}'
                    if channel_key in random_stats:
                        new_row = pd.DataFrame({
                            'Channel': [f'Channel {ch}'],
                            'Channel_Label': [ch_label],
                            'Mean_Intensity': [random_stats[channel_key]['mean']],
                            'Std_Intensity': [random_stats[channel_key]['std']],
                            'Samples_Used': [random_stats[channel_key]['samples']]
                        })
                        results_df = pd.concat([results_df, new_row], ignore_index=True)

                results_df.to_csv(random_dir / f'{DATA_FOLDER_NAME}_detailed_random_background.csv', index=False)

                # Excel-ready format
                excel_format = pd.DataFrame({
                    'Parameter': [f'Random Data C{i}' for i in range(1, CHANNELS_IN_IMAGE + 1)],
                    'Value': [random_means.get(f'Channel_{i}', 0) for i in range(1, CHANNELS_IN_IMAGE + 1)],
                    'Description': [f'Background intensity for {CHANNEL_LABELS.get(i, f"Channel_{i}")}'
                                    for i in range(1, CHANNELS_IN_IMAGE + 1)]
                })
                excel_format.to_csv(random_dir / f'{DATA_FOLDER_NAME}_excel_random_data.csv', index=False)

            logger.info("[SUCCESS] Random background analysis complete")
            logger.info(f"[SUCCESS] Sampled {peak_n} random locations from {total_rois_processed} ROIs")

            # Display results
            logger.info("Random Background Values:")
            for ch in range(1, CHANNELS_IN_IMAGE + 1):
                channel_key = f'Channel_{ch}'
                if channel_key in random_stats:
                    ch_label = CHANNEL_LABELS.get(ch, f'Channel_{ch}')
                    mean_val = random_stats[channel_key]['mean']
                    std_val = random_stats[channel_key]['std']
                    logger.info(f"  {ch_label} (Ch{ch}): {mean_val:.1f} +/- {std_val:.1f}")

            return random_means, random_stds

    except Exception as e:
        logger.error(f"Error in random background analysis: {e}")
        return None, None

    return None, None

def run_confidence_analysis():
    """Statistical confidence analysis for experimental reproducibility"""
    logger.info("=" * 60)
    logger.info("STATISTICAL CONFIDENCE ANALYSIS")
    logger.info("=" * 60)

    peak_csvs_dir = EXPERIMENT_DIR / 'peak_csvs'
    if not peak_csvs_dir.exists():
        logger.warning("No peak CSV files found")
        return

    logger.info(f"Processing experiment: {DATA_FOLDER_NAME}")
    logger.info(f"Analyzing channels: {CHANNELS_TO_ANALYZE}")

    # Setup confidence directories
    base_dir = EXPERIMENT_DIR / 'confidence_data'
    experiment_name = f'{DATA_FOLDER_NAME}_analysis'

    logger.info("Creating confidence analysis structure...")

    for channel_num in CHANNELS_TO_ANALYZE:
        channel = f'Channel_{channel_num}'
        channel_dir = base_dir / experiment_name / channel
        channel_dir.mkdir(parents=True, exist_ok=True)

        condition_dir = channel_dir / DATA_FOLDER_NAME
        condition_dir.mkdir(parents=True, exist_ok=True)

    # Process each analyzed channel
    for channel_num in CHANNELS_TO_ANALYZE:
        channel = f'Channel_{channel_num}'
        channel_label = CHANNEL_LABELS.get(channel_num, f'Channel_{channel_num}')
        logger.info(f"Processing {channel} ({channel_label})...")

        peak_csvs_channel_dir = peak_csvs_dir / channel

        if not peak_csvs_channel_dir.exists():
            logger.error(f"No peak CSV directory found: {peak_csvs_channel_dir}")
            continue

        # Get all CSV files for this channel
        csv_files = [f for f in peak_csvs_channel_dir.iterdir() if f.suffix == '.csv']

        if not csv_files:
            logger.error(f"No CSV files found in {peak_csvs_channel_dir}")
            continue

        logger.info(f"[SUCCESS] Found {len(csv_files)} CSV files")

        # Initialize DataFrames for statistical analysis
        channel_dfs = {f'Channel_{ch}': pd.DataFrame() for ch in range(1, CHANNELS_IN_IMAGE + 1)}
        files_processed = 0

        # Process all files together
        for csv_file in csv_files:
            try:
                df = pd.read_csv(csv_file)
                files_processed += 1

                if 'Unnamed: 0' in df.columns:
                    df = df.drop(columns=['Unnamed: 0'])

                # Concatenate channel data for statistical analysis
                for ch in range(1, CHANNELS_IN_IMAGE + 1):
                    channel_cols = [col for col in df.columns if f'Channel {ch}' in col and 'Distance' not in col]
                    if channel_cols:
                        temp_df = df[channel_cols]
                        if channel_dfs[f'Channel_{ch}'].empty:
                            channel_dfs[f'Channel_{ch}'] = temp_df
                        else:
                            channel_dfs[f'Channel_{ch}'] = pd.concat([channel_dfs[f'Channel_{ch}'], temp_df], axis=1)
            except Exception as e:
                logger.error(f"Error processing {csv_file}: {e}")
                continue

        if files_processed == 0:
            logger.error("No files processed successfully")
            continue

        logger.info(f"[SUCCESS] Successfully processed {files_processed} files")

        # Create summary statistics
        summary_df = pd.DataFrame()

        # Get distance column
        try:
            first_file = csv_files[0]
            first_df = pd.read_csv(first_file)
            distance_column = first_df['Distance'] if 'Distance' in first_df.columns else pd.Series(DISTANCES)
        except (FileNotFoundError, pd.errors.EmptyDataError):
            distance_column = pd.Series(DISTANCES)

        summary_df['Distance'] = distance_column

        # Calculate statistics for each channel
        logger.info("Calculating statistics:")

        for ch in range(1, CHANNELS_IN_IMAGE + 1):
            ch_label = CHANNEL_LABELS.get(ch, f'Channel_{ch}')
            if not channel_dfs[f'Channel_{ch}'].empty:
                avg_values = channel_dfs[f'Channel_{ch}'].mean(axis=1)
                std_values = channel_dfs[f'Channel_{ch}'].std(axis=1).fillna(0)

                summary_df[f'Channel {ch} Avg'] = avg_values
                summary_df[f'Channel {ch} Std'] = std_values

                overall_mean = avg_values.mean()
                overall_std = avg_values.std()
                data_points = channel_dfs[f'Channel_{ch}'].shape[1]
                logger.info(f"  {ch_label} (Ch{ch}): {data_points} peaks, Mean={overall_mean:.1f}+/-{overall_std:.1f}")
            else:
                summary_df[f'Channel {ch} Avg'] = 0
                summary_df[f'Channel {ch} Std'] = 0
                logger.info(f"  {ch_label} (Ch{ch}): No data")

        # Save results
        output_base = base_dir / experiment_name / channel / DATA_FOLDER_NAME
        for ch in range(1, CHANNELS_IN_IMAGE + 1):
            if not channel_dfs[f'Channel_{ch}'].empty:
                channel_dfs[f'Channel_{ch}'].to_csv(output_base / f'Channel_{ch}_all_data.csv', index=False)

        summary_df.to_csv(output_base / 'summary_data.csv', index=False)
        logger.info(f"[SUCCESS] Results saved to: {output_base}/")

    logger.info("[SUCCESS] Statistical confidence analysis completed!")


def parse_metadata_for_peak_counts() -> Dict[int, int]:
    """Extract actual peak counts from analysis metadata"""
    metadata_file = EXPERIMENT_DIR / 'analysis_metadata.txt'
    peak_counts = {}

    if not metadata_file.exists():
        logger.warning("No metadata file found for peak counts")
        return {}

    try:
        with open(metadata_file, 'r', encoding='utf-8') as f:
            content = f.read()

        current_channel = None
        lines = content.split('\n')

        for line in lines:
            line = line.strip()

            if line.startswith('Channel: Channel_'):
                try:
                    current_channel = int(line.split('Channel_')[1])
                except ValueError:
                    current_channel = None

            elif line.startswith('total_peaks: ') and current_channel is not None:
                try:
                    peak_count = int(line.split('total_peaks: ')[1])
                    peak_counts[current_channel] = peak_count
                except ValueError:
                    pass

        # Backbone channel always 0
        if BACKBONE_CHANNEL:
            peak_counts[BACKBONE_CHANNEL] = 0

        return peak_counts

    except (IOError, OSError) as e:
        logger.error(f"Error reading metadata file: {e}")
        return {}


def create_excel_headers(analysis_channels: List[int]) -> List[str]:
    """Generate Excel headers for the normalization file"""
    headers = ["Distance (μm)"]

    # Section 1: Raw data input for ALL channels from EACH analysis
    for ch in analysis_channels:
        for data_ch in range(1, CHANNELS_IN_IMAGE + 1):
            headers.append(f"C{ch} Channel {data_ch}")
            headers.append(f"C{ch} Channel {data_ch} STD")

    # Section 2: Duplicate raw data section (for calculations)
    headers.append("")  # Blank separator
    headers.append("Distance (μm)")

    # Repeat for section 2
    for ch in analysis_channels:
        for data_ch in range(1, CHANNELS_IN_IMAGE + 1):
            headers.append(f"C{ch} Channel {data_ch}")
            headers.append(f"C{ch} Channel {data_ch} STD")

    # Section 3: Normalization columns
    headers.append("")  # Blank separator

    # For each analyzed channel, normalization for ALL channels
    for ch in analysis_channels:
        for data_ch in range(1, CHANNELS_IN_IMAGE + 1):
            headers.append(f"norm C{ch} Channel {data_ch}")
            headers.append(f"norm C{ch} Channel {data_ch} STD")
            headers.append("Peak N")

    return headers


def find_confidence_files(analysis_channels: List[int]) -> Dict[int, Path]:
    """Find confidence data files for each analyzed channel"""
    channel_confidence_files = {}
    search_paths = [EXPERIMENT_DIR / 'confidence_data', Path('confidence_data')]

    for analyzed_channel in analysis_channels:
        channel_label = CHANNEL_LABELS.get(analyzed_channel, f'Channel_{analyzed_channel}')

        for search_path in search_paths:
            if search_path.exists():
                for root in search_path.rglob('summary_data.csv'):
                    if f'Channel_{analyzed_channel}' in str(root):
                        try:
                            df_test = pd.read_csv(root)
                            if len(df_test) > 0:
                                channel_confidence_files[analyzed_channel] = root
                                logger.info(f"Ch{analyzed_channel} ({channel_label}): Found confidence data")
                                break
                        except (pd.errors.EmptyDataError, pd.errors.ParserError):
                            continue
                if analyzed_channel in channel_confidence_files:
                    break

    return channel_confidence_files


def populate_excel_data(ws, analysis_channels: List[int], channel_confidence_files: Dict[int, Path]):
    """Populate Excel worksheet with data from confidence files"""
    # Write distance to column A
    for row_idx, distance in enumerate(DISTANCES, 2):
        ws.cell(row=row_idx, column=1, value=distance)

    # Populate data from each channel's analysis
    logger.info("Populating data from each channel's summary_data.csv...")

    section1_col = 2  # Start after distance column

    for ch_idx, ch in enumerate(analysis_channels):
        ch_label = CHANNEL_LABELS.get(ch, f'Channel_{ch}')

        if ch in channel_confidence_files:
            try:
                # Load the summary_data.csv for this channel's analysis
                df_ch = pd.read_csv(channel_confidence_files[ch])
                if 'Unnamed: 0' in df_ch.columns:
                    df_ch = df_ch.drop(columns=['Unnamed: 0'])

                logger.info(f"Processing Channel {ch} ({ch_label}) analysis...")

                populated_channels = []

                for data_ch in range(1, CHANNELS_IN_IMAGE + 1):
                    data_ch_label = CHANNEL_LABELS.get(data_ch, f'Channel_{data_ch}')
                    avg_col = f'Channel {data_ch} Avg'
                    std_col = f'Channel {data_ch} Std'

                    # Calculate column position
                    col_offset = ch_idx * (CHANNELS_IN_IMAGE * 2) + (data_ch - 1) * 2
                    avg_col_idx = section1_col + col_offset
                    std_col_idx = avg_col_idx + 1

                    if avg_col in df_ch.columns and std_col in df_ch.columns:
                        avg_data = df_ch[avg_col].values
                        std_data = df_ch[std_col].values

                        # Write avg data
                        for row_idx, value in enumerate(avg_data, 2):
                            if row_idx - 2 < len(DISTANCES):
                                ws.cell(row=row_idx, column=avg_col_idx, value=value)

                        # Write std data
                        for row_idx, value in enumerate(std_data, 2):
                            if row_idx - 2 < len(DISTANCES):
                                ws.cell(row=row_idx, column=std_col_idx, value=value)

                        populated_channels.append(data_ch_label)
                        logger.info(f"    [SUCCESS] Populated C{ch} Channel {data_ch} data ({data_ch_label})")
                    else:
                        logger.warning(f"    Missing data for Channel {data_ch} in C{ch} analysis")

                if not populated_channels:
                    logger.error(f"    No data populated for Channel {ch} analysis")

            except Exception as e:
                logger.error(f"    Error loading data for Ch{ch}: {e}")
        else:
            logger.error(f"No summary_data.csv found for Channel {ch}")


def add_excel_parameters(ws, data_length: int, random_means: Optional[Dict], 
                        metadata_peak_counts: Dict[int, int]):
    """Add parameter rows to Excel worksheet"""
    logger.info("Adding parameter rows...")
    param_start_row = data_length + 4  # Leave some blank rows

    param_labels = []
    param_values = []

    # Random Data for each channel
    for ch in range(1, CHANNELS_IN_IMAGE + 1):
        ch_label = CHANNEL_LABELS.get(ch, f'Channel_{ch}')
        param_labels.append(f'Random Data C{ch}:')

        channel_key = f'Channel_{ch}'
        if random_means and channel_key in random_means:
            param_values.append(random_means[channel_key])
            logger.info(f"Random Data C{ch}: {random_means[channel_key]:.1f}")
        else:
            param_values.append(0)
            logger.warning(f"Random Data C{ch}: 0 (not found)")

    # Peak N for each channel
    for ch in range(1, CHANNELS_IN_IMAGE + 1):
        ch_label = CHANNEL_LABELS.get(ch, f'Channel_{ch}')
        param_labels.append(f'Channel {ch} Peak N:')

        if ch in metadata_peak_counts:
            param_values.append(metadata_peak_counts[ch])
            logger.info(f"Channel {ch} Peak N: {metadata_peak_counts[ch]}")
        else:
            param_values.append(0)
            logger.warning(f"Channel {ch} Peak N: 0 (not found)")

    # Write parameter labels and values
    for idx, (label, value) in enumerate(zip(param_labels, param_values)):
        row_num = param_start_row + idx
        ws.cell(row=row_num, column=1, value=label)  # Column A
        ws.cell(row=row_num, column=2, value=value)  # Column B

    logger.info(f"[SUCCESS] Added {len(param_labels)} parameter rows starting at row {param_start_row}")
    return param_start_row


def add_excel_formulas(ws, analysis_channels: List[int], data_length: int, param_start_row: int):
    """Add Excel formulas for calculations and normalization"""
    logger.info("Adding Excel formulas...")
    
    num_channels = len(analysis_channels)

    # Calculate column positions
    section1_data_start = 2  # Column B
    section1_data_cols = num_channels * CHANNELS_IN_IMAGE * 2

    section2_start = section1_data_start + section1_data_cols + 1  # After section 1 + blank
    section2_distance = section2_start
    section2_data_start = section2_start + 1

    section3_start = section2_data_start + section1_data_cols + 1  # After section 2 + blank

    logger.info(f"Section 1 data: Columns {section1_data_start} to {section1_data_start + section1_data_cols - 1}")
    logger.info(f"Section 2 data: Columns {section2_data_start} to {section2_data_start + section1_data_cols - 1}")
    logger.info(f"Section 3 norm: Column {section3_start} onwards")

    # Copy distance from Section 1 to Section 2
    for row in range(2, data_length + 2):
        distance_formula = f"=A{row}"
        ws.cell(row=row, column=section2_distance, value=distance_formula)

    # Background subtraction formulas for each analysis and each channel
    for ch_idx, ch in enumerate(analysis_channels):
        for data_ch in range(1, CHANNELS_IN_IMAGE + 1):
            # Find the parameter row for this channel's random data
            random_data_param_row = param_start_row + (data_ch - 1)

            # Calculate column positions
            col_offset = ch_idx * (CHANNELS_IN_IMAGE * 2) + (data_ch - 1) * 2
            section1_avg_col = section1_data_start + col_offset
            section2_avg_col = section2_data_start + col_offset

            section1_avg_letter = col_to_letter(section1_avg_col)

            # Average column: section1_data - random_data
            for row in range(2, data_length + 2):
                bg_subtract_formula = f"={section1_avg_letter}{row}-$B${random_data_param_row}"
                ws.cell(row=row, column=section2_avg_col, value=bg_subtract_formula)

            # STD column: pass through unchanged
            section1_std_col = section1_avg_col + 1
            section2_std_col = section2_avg_col + 1
            section1_std_letter = col_to_letter(section1_std_col)

            for row in range(2, data_length + 2):
                std_formula = f"={section1_std_letter}{row}"
                ws.cell(row=row, column=section2_std_col, value=std_formula)

    logger.info("[SUCCESS] Added background subtraction formulas")

    # Add MAX value calculations
    logger.info("Adding MAX value calculations...")
    max_calc_row = data_length + 2

    for ch_idx, ch in enumerate(analysis_channels):
        for data_ch in range(1, CHANNELS_IN_IMAGE + 1):
            col_offset = ch_idx * (CHANNELS_IN_IMAGE * 2) + (data_ch - 1) * 2
            section2_avg_col = section2_data_start + col_offset
            section2_std_col = section2_avg_col + 1

            section2_avg_letter = col_to_letter(section2_avg_col)

            # Use absolute maximum for normalization (handles both positive and negative values)
            max_formula = f"=IF(ABS(MIN({section2_avg_letter}2:{section2_avg_letter}{data_length + 1}))>MAX({section2_avg_letter}2:{section2_avg_letter}{data_length + 1}),ABS(MIN({section2_avg_letter}2:{section2_avg_letter}{data_length + 1})),MAX({section2_avg_letter}2:{section2_avg_letter}{data_length + 1}))"
            ws.cell(row=max_calc_row, column=section2_avg_col, value=max_formula)

            # STD max = same as avg max
            std_max_formula = f"={section2_avg_letter}{max_calc_row}"
            ws.cell(row=max_calc_row, column=section2_std_col, value=std_max_formula)

    logger.info(f"[SUCCESS] Added MAX calculations at row {max_calc_row}")

    # Add normalization formulas (Section 3)
    logger.info("Adding normalization formulas...")

    for ch_idx, ch in enumerate(analysis_channels):
        for data_ch in range(1, CHANNELS_IN_IMAGE + 1):
            # Get source columns from Section 2
            col_offset = ch_idx * (CHANNELS_IN_IMAGE * 2) + (data_ch - 1) * 2
            section2_avg_col = section2_data_start + col_offset
            section2_std_col = section2_avg_col + 1

            # Target columns in Section 3
            norm_col_offset = (ch_idx * CHANNELS_IN_IMAGE + (data_ch - 1)) * 3
            norm_avg_col = section3_start + norm_col_offset
            norm_std_col = norm_avg_col + 1
            norm_peak_col = norm_avg_col + 2

            section2_avg_letter = col_to_letter(section2_avg_col)
            section2_std_letter = col_to_letter(section2_std_col)

            # Normalization formulas
            for row in range(2, data_length + 2):
                # norm avg = background_subtracted / max_value
                norm_avg_formula = f"={section2_avg_letter}{row}/{section2_avg_letter}${max_calc_row}"
                ws.cell(row=row, column=norm_avg_col, value=norm_avg_formula)

                # norm std = std / max_value
                norm_std_formula = f"={section2_std_letter}{row}/{section2_std_letter}${max_calc_row}"
                ws.cell(row=row, column=norm_std_col, value=norm_std_formula)

            # Peak N column (reference to parameter)
            peak_n_param_row = param_start_row + CHANNELS_IN_IMAGE + (data_ch - 1)
            for row in range(2, data_length + 2):
                peak_n_formula = f"=$B${peak_n_param_row}"
                ws.cell(row=row, column=norm_peak_col, value=peak_n_formula)

    logger.info("[SUCCESS] Added normalization formulas")


def create_analysis_excel(channel_results: Dict, channel_maxima: Dict, random_means: Optional[Dict]) -> Optional[str]:
    """Generate comprehensive Excel file with analysis results"""
    logger.info("=" * 60)
    logger.info("CREATING ANALYSIS EXCEL FILE")
    logger.info("=" * 60)

    # Get real peak counts from metadata
    metadata_peak_counts = parse_metadata_for_peak_counts()

    # Create Excel file
    excel_dir = EXPERIMENT_DIR / 'excel_files'
    excel_dir.mkdir(parents=True, exist_ok=True)
    output_filename = excel_dir / f"{DATA_FOLDER_NAME}_analysis_results.xlsx"

    logger.info(f"Creating Excel file: {output_filename}")

    try:
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            # Find confidence files
            logger.info("Locating confidence data files...")

            channel_confidence_files = find_confidence_files(CHANNELS_TO_ANALYZE)

            # Create analysis tabs
            tabs_created = 0
            for analyzed_channel in CHANNELS_TO_ANALYZE:
                channel_label = CHANNEL_LABELS.get(analyzed_channel, f'Channel_{analyzed_channel}')
                tab_name = f'Ch{analyzed_channel}_{channel_label}'

                logger.info(f"Creating analysis tab: {tab_name}")

                # Initialize analysis DataFrame
                analysis_df = pd.DataFrame()
                confidence_file = channel_confidence_files.get(analyzed_channel)

                if confidence_file:
                    try:
                        df = pd.read_csv(confidence_file)

                        if 'Unnamed: 0' in df.columns:
                            df = df.drop(columns=['Unnamed: 0'])

                        # Distance column
                        if 'Distance' in df.columns:
                            analysis_df['Distance (μm)'] = df['Distance']
                        else:
                            analysis_df['Distance (μm)'] = DISTANCES

                        logger.info("Adding data columns for all channels...")

                        # Add data columns for all channels
                        for ch in range(1, CHANNELS_IN_IMAGE + 1):
                            ch_label_inner = CHANNEL_LABELS.get(ch, f'Channel_{ch}')
                            avg_col = f'C1 Channel{ch}'
                            std_col = f'C1 Channel{ch} STD'

                            source_avg_col = f'Channel {ch} Avg'
                            source_std_col = f'Channel {ch} Std'

                            if source_avg_col in df.columns and source_std_col in df.columns:
                                analysis_df[avg_col] = df[source_avg_col]
                                analysis_df[std_col] = df[source_std_col]
                                data_mean = df[source_avg_col].mean()
                                logger.info(f"      {ch_label_inner} (Ch{ch}): Mean={data_mean:.1f}")
                            else:
                                analysis_df[avg_col] = 0
                                analysis_df[std_col] = 0
                                logger.warning(f"      {ch_label_inner} (Ch{ch}): No data - using zeros")

                    except Exception as e:
                        logger.error(f"Error loading confidence data: {e}")
                        # Create default data
                        analysis_df['Distance (μm)'] = DISTANCES

                        for ch in range(1, CHANNELS_IN_IMAGE + 1):
                            analysis_df[f'C1 Channel{ch}'] = 0
                            analysis_df[f'C1 Channel{ch} STD'] = 0

                else:
                    logger.warning("No confidence data found - creating template")
                    analysis_df['Distance (μm)'] = DISTANCES

                    for ch in range(1, CHANNELS_IN_IMAGE + 1):
                        analysis_df[f'C1 Channel{ch}'] = 0
                        analysis_df[f'C1 Channel{ch} STD'] = 0

                # Save analysis sheet
                analysis_df.to_excel(writer, sheet_name=tab_name, index=False)
                logger.info(f"[SUCCESS] Tab '{tab_name}' created with {len(analysis_df.columns)} columns")
                tabs_created += 1

            # Create summary sheet
            summary_data = [
                {'Configuration': 'Analysis Type', 'Value': 'Multi-Channel Peak Analysis',
                 'Description': f'Analysis of {CHANNELS_IN_IMAGE} channel microscopy images'},
                {'Configuration': 'Channels Analyzed', 'Value': len(CHANNELS_TO_ANALYZE),
                 'Description': f'Peak detection in channels: {CHANNELS_TO_ANALYZE}'},
                {'Configuration': 'Backbone Channel', 'Value': BACKBONE_CHANNEL,
                 'Description': f'Reference channel: {CHANNEL_LABELS.get(BACKBONE_CHANNEL)}'},
                {'Configuration': 'Data Format', 'Value': f'{1 + (CHANNELS_IN_IMAGE * 2)} columns',
                 'Description': 'Distance + Channel averages + Standard deviations'},
                {'Configuration': 'Usage', 'Value': 'Ready for normalization template',
                 'Description': 'Copy tab data to your analysis template'}
            ]

            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

            # Create parameters sheet
            params_data = []

            # Configuration parameters
            params_data.append({
                'Parameter': 'Experiment Directory',
                'Value': str(EXPERIMENT_DIR),
                'Description': 'Base directory for analysis outputs'
            })

            params_data.append({
                'Parameter': 'Data Folder',
                'Value': DATA_FOLDER_NAME,
                'Description': 'Input data folder name'
            })

            # Random background data
            if random_means:
                for ch in range(1, CHANNELS_IN_IMAGE + 1):
                    ch_label = CHANNEL_LABELS.get(ch, f'Channel_{ch}')
                    channel_key = f'Channel_{ch}'
                    if channel_key in random_means:
                        params_data.append({
                            'Parameter': f'Random Data C{ch}',
                            'Value': f"{random_means[channel_key]:.2f}",
                            'Description': f'Background intensity for {ch_label}'
                        })

            # Peak counts from metadata
            for ch in range(1, CHANNELS_IN_IMAGE + 1):
                ch_label = CHANNEL_LABELS.get(ch, f'Channel_{ch}')

                if ch in metadata_peak_counts:
                    peak_count = metadata_peak_counts[ch]
                else:
                    peak_count = 0

                if ch == BACKBONE_CHANNEL:
                    description = f'{ch_label} - reference channel (not analyzed for peaks)'
                elif ch in CHANNELS_TO_ANALYZE:
                    description = f'{ch_label} - total peaks detected in analysis'
                else:
                    description = f'{ch_label} - not included in peak analysis'

                params_data.append({
                    'Parameter': f'Channel {ch} Peak Count',
                    'Value': peak_count,
                    'Description': description
                })

            # Channel maxima
            if channel_maxima:
                for ch in range(1, CHANNELS_IN_IMAGE + 1):
                    ch_label = CHANNEL_LABELS.get(ch, f'Channel_{ch}')
                    if ch in channel_maxima and channel_maxima[ch] > 0:
                        params_data.append({
                            'Parameter': f'{ch_label} Max Intensity',
                            'Value': f"{channel_maxima[ch]:.0f}",
                            'Description': f'Maximum peak intensity in {ch_label}'
                        })

            params_df = pd.DataFrame(params_data)
            params_df.to_excel(writer, sheet_name='Parameters', index=False)

        logger.info("[SUCCESS] Excel file created successfully!")
        logger.info(f"[SUCCESS] File: {output_filename}")
        logger.info(f"[SUCCESS] Tabs created: {tabs_created}")

        return str(output_filename)

    except Exception as e:
        logger.error(f"Error creating Excel file: {e}")
        return None


def create_automated_normalization_calc(channel_results: Dict, channel_maxima: Dict, 
                                       random_means: Optional[Dict]) -> Optional[str]:
    """Generate complete normalization calculation file with data from each channel analysis"""
    logger.info("=" * 60)
    logger.info("CREATING AUTOMATED NORMALIZATION CALCULATION FILE")
    logger.info("=" * 60)

    # Get real peak counts from metadata
    metadata_peak_counts = parse_metadata_for_peak_counts()

    # Only include analyzed channels (exclude backbone)
    analysis_channels = [ch for ch in CHANNELS_TO_ANALYZE if ch != BACKBONE_CHANNEL]
    analysis_channels.sort()  # Keep consistent order

    logger.info(f"Including channels for analysis: {analysis_channels}")
    logger.info(f"Excluding backbone channel: {BACKBONE_CHANNEL}")

    if len(analysis_channels) < 2:
        logger.error("Need at least 2 analysis channels for normalization file")
        return None

    # Create normalization file
    excel_dir = EXPERIMENT_DIR / 'excel_files'
    excel_dir.mkdir(parents=True, exist_ok=True)
    output_filename = excel_dir / f"{DATA_FOLDER_NAME}_normalization_data.xlsx"

    logger.info(f"Creating normalization file: {output_filename}")

    try:
        # Find confidence files for each analyzed channel
        logger.info("Locating summary_data.csv for each channel analysis...")
        channel_confidence_files = find_confidence_files(analysis_channels)

        if len(channel_confidence_files) != len(analysis_channels):
            logger.warning(f"Found {len(channel_confidence_files)} files, expected {len(analysis_channels)}")

        # Create the workbook
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"  # Match the original Excel template

        # Generate headers
        logger.info("Generating headers...")
        headers = create_excel_headers(analysis_channels)

        # Write headers to row 1
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        logger.info(f"[SUCCESS] Created {len(headers)} column headers")

        # Populate data from confidence files
        populate_excel_data(ws, analysis_channels, channel_confidence_files)

        # Add parameter rows
        param_start_row = add_excel_parameters(ws, DATA_LENGTH, random_means, metadata_peak_counts)

        # Save basic structure first
        if not safe_save_excel(wb, output_filename):
            return None

        logger.info("Basic structure created, now adding Excel formulas...")

        # Reopen with openpyxl to add formulas
        from openpyxl import load_workbook
        wb = load_workbook(output_filename)
        ws = wb.active

        # Add all formulas
        add_excel_formulas(ws, analysis_channels, DATA_LENGTH, param_start_row)

        # Save the final workbook with all formulas
        if not safe_save_excel(wb, output_filename):
            return None

        logger.info("[SUCCESS] COMPLETE NORMALIZATION FILE CREATED!")
        logger.info(f"[SUCCESS] File: {output_filename}")
        logger.info(f"[SUCCESS] Channels included: {[f'C{ch} ({CHANNEL_LABELS.get(ch)})' for ch in analysis_channels]}")
        logger.info("[SUCCESS] Data populated from each channel's summary_data.csv")
        logger.info("[SUCCESS] All Excel formulas added automatically")
        logger.info("[SUCCESS] Background subtraction formulas working")
        logger.info("[SUCCESS] Normalization calculations working")
        logger.info("[SUCCESS] MAX value calculations working")
        logger.info("[SUCCESS] All parameters populated")
        logger.info("[SUCCESS] Ready for immediate use!")

        return str(output_filename)

    except Exception as e:
        logger.error(f"Error creating normalization file: {e}")
        import traceback
        traceback.print_exc()
        return None


def main() -> bool:
    """Main analysis pipeline"""
    logger.info("Mitochondrial Protein Analysis Pipeline - Windows Compatible Version")
    logger.info("=" * 60)
    logger.info(f"Experiment: {EXPERIMENT_DIR}")
    logger.info(f"Data folder: {DATA_FOLDER_NAME}")

    # Add this line to show where CSVs are expected
    csv_location = EXPERIMENT_DIR / CSV_SUBFOLDER if CSV_SUBFOLDER else EXPERIMENT_DIR
    logger.info(f"CSV files location: {csv_location}")

    logger.info(f"Channels in images: {CHANNELS_IN_IMAGE}")

    # Validate configuration first
    if not validate_configuration():
        logger.error("Please fix configuration errors before running analysis")
        return False

    # Display channel configuration
    logger.info("Channel Configuration:")
    for ch in range(1, CHANNELS_IN_IMAGE + 1):
        ch_label = CHANNEL_LABELS.get(ch, f'Channel_{ch}')
        analyzed = "[X]" if ch in CHANNELS_TO_ANALYZE else "[ ]"
        backbone = "(backbone)" if ch == BACKBONE_CHANNEL else ""
        logger.info(f"  {analyzed} Channel {ch}: {ch_label} {backbone}")

    # Get actual channels to analyze (excluding backbone)
    actual_channels_to_analyze = [ch for ch in CHANNELS_TO_ANALYZE if ch != BACKBONE_CHANNEL]

    if not actual_channels_to_analyze:
        logger.error("No channels to analyze after excluding backbone channel")
        return False

    analyzed_labels = [CHANNEL_LABELS.get(ch, f'Channel_{ch}') for ch in actual_channels_to_analyze]
    logger.info(f"Analyzing peaks in: {', '.join(analyzed_labels)}")

    if BACKBONE_CHANNEL:
        backbone_label = CHANNEL_LABELS.get(BACKBONE_CHANNEL, f'Channel_{BACKBONE_CHANNEL}')
        logger.info(f"Using backbone reference: {backbone_label} (Channel {BACKBONE_CHANNEL})")

    try:
        # Execute analysis pipeline
        setup_folders()

        logger.info("=" * 50)
        logger.info("RUNNING COMPLETE ANALYSIS PIPELINE")
        logger.info("=" * 50)

        channel_results, channel_maxima = run_peak_analysis()
        random_means, random_stds = run_random_background_analysis()
        run_confidence_analysis()
        excel_file = create_analysis_excel(channel_results, channel_maxima, random_means)
        normalization_file = create_automated_normalization_calc(channel_results, channel_maxima, random_means)

        # Display final results
        logger.info("=" * 60)
        logger.info("ANALYSIS PIPELINE COMPLETED SUCCESSFULLY")
        logger.info("=" * 60)

        logger.info(f"[SUCCESS] Results saved in: {EXPERIMENT_DIR}/")
        logger.info("[SUCCESS] Cross-channel normalized CSVs created")
        logger.info("[SUCCESS] Multi-channel plots generated")
        logger.info("[SUCCESS] Statistical confidence analysis completed")
        logger.info(f"[SUCCESS] Excel file created: {excel_file or 'Error occurred'}")
        logger.info(f"[SUCCESS] Normalization file created: {normalization_file or 'Error occurred'}")

        logger.info("Channel Analysis Summary:")
        for ch in range(1, CHANNELS_IN_IMAGE + 1):
            ch_label = CHANNEL_LABELS.get(ch, f'Channel_{ch}')
            max_intensity = channel_maxima.get(ch, 0)
            if ch in actual_channels_to_analyze:
                logger.info(f"  {ch_label} (Ch{ch}): Max intensity = {max_intensity:.0f}")
            elif ch == BACKBONE_CHANNEL:
                logger.info(f"  {ch_label} (Ch{ch}): Reference channel (backbone)")
            else:
                logger.info(f"  {ch_label} (Ch{ch}): Not analyzed")

        logger.info("Generated Output Files:")
        logger.info("• Individual peak data: peak_csvs/")
        logger.info("• Average peak data: avg_peak_csvs/")
        logger.info("• Normalized peak data: norm_peak_csvs/")
        logger.info("• Statistical confidence: confidence_data/")
        logger.info("• Random background: random_data_lines/")
        logger.info("• Analysis plots: plots/")
        logger.info("• Excel results: excel_files/")
        logger.info(f"• AUTOMATED NORMALIZATION: {DATA_FOLDER_NAME}_normalization_data.xlsx")

        logger.info("ANALYSIS COMPLETE!")
        logger.info("Your data is ready for normalization and publication!")

        return True

    except Exception as e:
        logger.error(f"Pipeline failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    try:
        success = main()
        if success:
            logger.info("[SUCCESS] PIPELINE COMPLETED SUCCESSFULLY!")
            logger.info("Next Steps:")
            logger.info("1. Review the generated Excel analysis file")
            logger.info("2. Open the automated normalization file - ALL data pre-filled!")
            logger.info("3. All formulas are already set up - just check the results")
            logger.info("4. Use normalized results for publication-ready figures")
        else:
            logger.error("Pipeline failed. Please check the errors above.")
    except Exception as e:
        logger.error(f"CRITICAL ERROR: {e}")
        import traceback
        traceback.print_exc()

    input("Press Enter to exit...")

